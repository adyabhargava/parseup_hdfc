import time
import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver
import constants
from openpyxl.styles import PatternFill


class backEnd:
    """ This class contains the backend code of the method."""

    def __init__(self, file, sheet, address, column_to_extract, column_to_insert, organization):
        """ This method initializes the instance variables.
        It is used to link the frontend inputs to the backend."""
        self.filePath = file
        self.sheetName = sheet
        self.organization = organization
        self.websites = address
        self.column_to_append = column_to_insert
        self.column_to_extract = column_to_extract
        #Dictionary with key as the name of item and value as the price.
        """self.names is for items written in the excel sheet"""
        self.names = {}
        self.items = []
        self.prices = []
        """self.all_website_items is for items on the web. This is cross checked with self.names"""
        self.all_website_items = {}
        """This stores all the new items, found on comparing self.names with self.all_website_items"""
        self.new = []
        """This stores all the new items, found on comparing self.items with self.names"""
        self.non_existent = []


    def extract_from_excel(self):
        """This method extracts data from the given excel file."""
        #Opens the Excel File
        self.ps = openpyxl.load_workbook(self.filePath)
        #Opens the Sheet in the excel file
        self.sheet = self.ps[self.sheetName]
        #Loops over each row in the column to parse to get the data from it.
        #Starts looping from row 2.
        for row in range(2, self.sheet.max_row):
            # each row in the spreadsheet represents information for a particular item.
            item = self.sheet[self.column_to_extract + str(row)].value
            if item != None:
                #It adds that item to self.names dictionary with value being empty string right now. Price is later
                #added to this.
                self.names[item] = ""


    def parse_website(self, website):
        #IMPORTANT!!: Need to find a way to find location of chromedriver
        self.driver = webdriver.Chrome(constants.chromedriver_location)
        self.driver.get(website)
        time.sleep(5)
        self.html = self.driver.page_source
        self.soup = BeautifulSoup(self.html, "html.parser")



    def no_double_count(self, all_items):
        """
        This is common to all web scrapers
        Since we loop over each page which may repeat items across page, self.all_website_items may already
        contain the item name.
        This method checks if the self.all_website_items already has the given item in it.
        If the item is not present in the list, only then does it add it to the dictionary self.all_web_items;
        otherwise, not."""
        for item in all_items:
            #Since we loop over each page which may repeat items across page, self.all_website_items may already
            #contain the item name
            if not self.all_website_items.__contains__(item):
                #To the field all_website_names, this adds an item if it is not already present.
                to_add = item.text
                #if item.text.__contains__("\t"):
                #to_add = item.text[:len(item.text)-1]x
                self.all_website_items[to_add] = ""


    def extract_prices_mcd(self, item_name):
        """This extracts & scrapes the price of the given item in the function and puts it against the name
        of the item in the dictionary. This only gives way to get the price of a particlar item"""
        """ This checks a particular web link for an item name. It that item is found on that webpage, it stores its
        price in a dictionary corresponding to the item name. If it is not found, (and has not been found before on any other
        webpage, it puts its value as 'Not Available' in self.names. Else, it does nothing."""
        #Finds whether or not that element is there in self.soup
        self.elem = self.soup.find("div", text = item_name)
        #print(item_name + "has self.elem")
        #print(self.elem)
        #If it is found, it does the following:
        if self.elem:
            #Adds its price corresponding to its name in self.names
            self.all_website_items[item_name] = self.elem.parent.parent.find("span",{"class":"price pr-1"}).text
            #Adds item name to the list of items whose prices have already been found.
            self.items.append(item_name)


    def extract_prices_bk(self, item):
        """Extracts prices from Burger King's HTML. This only gives way to get the price of a particlar item"""

    def extract_prices_dominos(self, item):
        """Extracts prices from Dominos' HTML. This only gives way to get the price of a particlar item"""

    def extract_prices_kfc(self, item):
        """Extracts prices from KFC's HTML. This only gives way to get the price of a particlar item"""

    def extract_prices_pizzahut(self, item):
        """Extracts prices from Pizza Hut's HTML. This only gives way to get the price of a particlar item"""


    def get_webitems_mcd(self):
        """ This lists down all the elements from mcdonalds and stores it in element self.all_website_items"""
        #Scaps website to find all items.
        elem = self.soup.find("div", id = "home-page-wrapper")
        #all_items is a list with all the items' html code format.
        all_items = elem.findAll('div', {"class": "item-title"})
        #self.all_website_items thus is a list of all items on a menu.
        self.no_double_count(all_items)

    def get_webitems_bk(self):
        """Extracts prices from Burger King's HTML"""

    def get_webitems_dominos(self):
        """Extracts prices from Dominos' HTML"""

    def get_webitems_kfc(self):
        """Extracts prices from KFC's HTML"""

    def get_webitems_pizzahut(self):
        """Extracts prices from Pizza Hut's HTML"""

    def get_all_web_items(self):
        """This is the main function that calls on specific helper methods & gets all the web items from a particular page"""
        if self.organization == "McDonalds":
            self.get_webitems_mcd()
        if self.organization == "KFC":
            self.get_webitems_kfc()
        if self.organization == "Burger King":
            self.get_webitems_bk()
        if self.organization == "Pizza Hut":
            self.get_webitems_pizzahut()
        if self.organization == "Dominos":
            self.get_webitems_dominos()

    def extract_price(self, item):
        #TO BE PUT IN BACKEND
        """This calls respective helper functions based on the name of the org."""
        if self.organization == "McDonalds":
            self.extract_prices_mcd(item)
        if self.organization == "KFC":
            self.extract_prices_kfc(item)
        if self.organization == "Burger King":
            self.extract_prices_bk(item)
        if self.organization == "Pizza Hut":
            self.extract_prices_pizzahut(item)
        if self.organization == "Dominos":
            self.extract_prices_dominos(item)


    def finds_new(self):
        """This finds a list of objects that are present in the new menu but not in the old menu"""
        """Checks if there are any new items. Returns True if there are, False if there are not. Updates self.new"""
        for item in self.items:
            if not self.names.__contains__(item):
                self.new.append(item)
        if len(self.new) == 0:
            return False
        else:
            return True

    def put_into_excel(self):
        """This method ensures that all the known prices are put into the excel sheet corresponding to the item names.
        It works the following way: it loops over each element in self.names, searches it up in self.all_website_items"""
        for row in range(1, self.sheet.max_row+2):
            # each row in the spreadsheet represents information for a particular item.
            item = self.sheet[self.column_to_extract + str(row)].value
            #It adds that item to self.names dictionary with value being empty string right now. Price is later
            #added to this.
            if self.items.__contains__(item):
                price = self.all_website_items[item]
                cell = self.column_to_append+str(row)
                self.sheet[cell] = price
                self.ps.save(self.filePath)
            else:
                price = "Not Available"
                cell = self.column_to_append+str(row)
                self.sheet[cell] = price
                cell_of_items = self.column_to_extract+str(row)
                cell_of_price = self.column_to_append+str(row)
                yellow = "00FFFF00"
                highlight = PatternFill(start_color=yellow,
                                        end_color=yellow,
                                        fill_type='solid')
                self.sheet[cell_of_items].fill = highlight
                self.sheet[cell_of_price].fill = highlight
                self.ps.save(self.filePath)


    def handle_new_items(self):
        """This starts adding the new items from self.new into the excel sheet, starting from the last row after
        the last item. It also highlights these."""
        starting_row = self.sheet.max_row+1
        for new in self.new:
            cell_of_items = self.column_to_extract+str(starting_row)
            self.sheet[cell_of_items] = new
            price = self.all_website_items[new]
            cell_of_price = self.column_to_append+str(starting_row)
            self.sheet[cell_of_price] = price
            yellow = "00FFFF00"
            highlight = PatternFill(start_color=yellow,
                                    end_color=yellow,
                                    fill_type='solid')
            self.sheet[cell_of_items].fill = highlight
            self.sheet[cell_of_price].fill = highlight
            self.ps.save(self.filePath)
            starting_row +=1


    def extract_all_info(self):
        #PUT IN BACKEND
        """Extracts all prices for a particular QSR company. This is main function to extract, it only calls
        a number of helper functions"""
        for section in self.websites:
            #Opens up website in order to parse it.
            self.parse_website(section)
            #Gets all items from that page- scrape it from the website. Then, add it to dictionary self.all_website_items
            #and list self.items.
            self.get_all_web_items()
            #This looks at each item in self.all_website_items and gets its price
            #After this, we know the price of each item, and for those items that we know price of, they are present in
            #the list self.items; Now we must first compare to items we don't know price of, and then compare to items
            #that are new, and highlight each of these yellow.
            for item in self.all_website_items:
                self.extract_price(item)
        self.put_into_excel()
        new = self.finds_new()
        if new:
            self.handle_new_items()
